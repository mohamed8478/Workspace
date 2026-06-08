from datetime import datetime, date, timedelta
import pyodbc
import pandas as pd
from sqlalchemy import create_engine
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import os
import html


# Output configuration
OUTPUT_FOLDER = os.path.dirname(os.path.abspath(__file__))  # Current script folder
OUTPUT_FILE = os.path.join(OUTPUT_FOLDER, "Sortie Plâtre par Client.xlsx")


# Category mapping for product classification
CATEGORY_MAPPING = {
    "SULFATE DE CALCIUM HEMIHYDRATE EN BIG BA": "AGRIGYPSE",
    "PLATRE GAZELLE EXTRA FIN EN PO": "PLATRE",
    "AGRIGYPSE 0-20 VRAC": "AGRIGYPSE",
    "AGRIGYPSE BIG BAG": "AGRIGYPSE",
    "AGRIGYPSE PP 50 KG": "AGRIGYPSE",
    "AGRIGYPSE EN VRAC": "AGRIGYPSE",
    "AGRIGYPSE PP": "AGRIGYPSE",
    "GYPSE BROYE AGRIGYPSE SAC 70KG": "AGRIGYPSE",
    "SULFATE DE CALCIUM": "AGRIGYPSE",
    "PLATRE EN VRAC": "AGRIGYPSE",
    "GYPSE BROYE": "AGRIGYPSE",
    "DALLE TOURBILLON": "DALLES",
    "DALLE SOLEIL SANS STRUCTURE": "DALLES",
    "DALLE SABLE SANS STRUCTURE": "DALLES",
    "DALLE MEDITERRANE": "DALLES",
    "DALLE HELICE": "DALLES",
    "DALLE FISSURE SANS STRUCTURE": "DALLES",
    "DALLE TOURBILLON DEPART SAFI AVEC STRUCTURE": "DALLES",
    "DALLE TOURBILLON DEPART SAFI SANS STRUCTURE": "DALLES",
    "DALLE SOLEIL DEPART SAFI AVEC STRUCTURE": "DALLES",
    "DALLE SOLEIL DEPART SAFI SANS STRUCTURE": "DALLES",
    "DALLE SABLE DEPART SAFI AVEC STRUCTURE": "DALLES",
    "DALLE SABLE DEPART SAFI SANS STRUCTURE": "DALLES",
    "DALLE PERFORE ROND DEPART SAFI SANS STRUCTURE": "DALLES",
    "DALLE PERFORE CARRE DEPART SAFI AVEC STRUCTURE": "DALLES",
    "DALLE PERFORE CARRE DEPART SAFI SANS STRUCTURE": "DALLES",
    "DALLE MEDITERRANE DEPART SAFI AVEC STRUCTURE": "DALLES",
    "DALLE MEDITERRANE DEPART SAFI SANS STRUCTURE": "DALLES",
    "DALLE HELICE DEPART SAFI AVEC STRUCTURE": "DALLES",
    "DALLE HELICE DEPART SAFI SANS STRUCTURE": "DALLES",
    "DALLE FISSURE DEPART SAFI AVEC STRUCTURE": "DALLES",
    "DALLE FISSURE DEPART SAFI SANS STRUCTURE": "DALLES",
    "AGRIGYPSE PP 50 KG EXPORT": "EXPORT",
    "WHITE GYPSUM ROCK 200-400MM": "EXPORT",
    "PLATRE EXPORT EXTRA FIN GAZELLE": "EXPORT",
    "PLATRE PGC EXPORT NEJMA": "EXPORT",
    "PLATRE EXPORT EXTRA FIN NEJMA": "EXPORT",
    "PLATRE MOULAGE GAZELLE EXPORT": "EXPORT",
    "GYPSE BLANC ROCHE EXPORT BI": "EXPORT",
    "PLATRE PGC EXPORT SAVANE": "EXPORT",
    "PLATRE PGC EXPORT GAZELLE": "EXPORT",
    "GYPSE BLANC CONCASSE EXPORT BI": "EXPORT",
    "PLATRE EXPORT SPECIAL THE DOVE": "EXPORT",
    "PLATRE EXPORT SPECIAL \"THE DOVE\"": "EXPORT",
    "GYPSE BLANC ROCHE EXPORT BIG BAG": "EXPORT",
    "GYPSE BLANC CONCASSE EXPORT BIG BAG": "EXPORT",
    "PLATRE EXPORT MOULAGE SAVANE": "EXPORT",
    "PLATRE EXPORT EN SAC EN BIG BAG": "EXPORT",
    "PLATRE EXPORT SPECIAL GAZELLE": "EXPORT",
    "PLATRE PGC EXPORT GAZELLE_": "EXPORT",
    "PLATRE EXPORT": "EXPORT",
    "ENDUIT JOINT EJ4": "MORTIER",
    "ENDUIT DE PEINTURE": "MORTIER",
    "ENDUIT EJ8": "MORTIER",
    "TALOCHE S9AF VERT PP": "MORTIER",
    "MORTIER DE PLATRE MPREMIUM PP": "MORTIER",
    "MORTIER DE PLATRE GAZELLE M.P.": "MORTIER",
    "MORTIER DE FINITION TOP  FINO P": "MORTIER",
    "MORTIER DE PLATRE GAZELLE TALO": "MORTIER",
    "ENDUIT EJ8 EN PALETTE": "MORTIER",
    "ENDUIT DE PEINTURE ECHANTILLON": "MORTIER",
    "ENDUIT JOINT EJ8 ECHANTILLON": "MORTIER",
    "ENDUIT JOINT EJ4 ech": "MORTIER",
    "MORTIER DE PLATRE MPREMIUM": "MORTIER",
    "ENDUIT DE FINITION": "MORTIER",
    "MORTIER DE PLATRE GAZELLE M.P.E 80 S SUPER": "MORTIER",
    "MORTIER DE PLATRE GAZELLE M.P.E 80 L": "MORTIER",
    "MORTIER EN PALETTE": "MORTIER",
    "MORTIER DE PLATRE GAZELLE TALOCHE 60": "MORTIER",
    "MORTIER MPE 80 BLEU PP": "MORTIER",
    "SPECIAL NEJMA": "PLATRE",
    "SPECIAL DOVE": "PLATRE",
    "SAVANE": "PLATRE",
    "PLATRE MOULAGE SPECIAL POLYPRO": "PLATRE",
    "PLATRE DE CONSTRUCTION ROUGE P": "PLATRE",
    "MOULDING NEJMA": "PLATRE",
    "PLATRE MOULAGE GAZELLE": "PLATRE",
    "EXTRA FIN INDUSTRIEL ECHANTILLON": "PLATRE",
    "PLATRE MOULAGE SPECIAL POLYPROPYLENE": "PLATRE",
    "PLATRE DE MOULAGE SPECIAL POLYPROPYLENE": "PLATRE",
    "PLATRE NEJMA EN POLYPROPYLENE": "PLATRE",
    "PLATRE DE CONSTRUCTION EN POLYPROPYLENE ROUGE": "PLATRE",
    "PLATRE DE CONSTRUCTION EN POLYPROPYLENE BLEU": "PLATRE",
    "PLATRE ( 2 TONNES)": "PLATRE",
    "PLATRE EN SACS POLYPROPYLENE ROUGE": "PLATRE",
    "PLATRE EN STOCK": "PLATRE",
    "PLATRE POLYPROPYLENE EN PALETTE": "PLATRE",
    "PLATRE MOULLAGE GAZELLE EN VRAC": "PLATRE",
    "PLATRE EN BIG BAG": "PLATRE",
    "PLATRE DE MOULAGE SPECIAL": "PLATRE",
    "PLATRE NEJMA": "PLATRE",
    "PLATRE DE CONSTRUCTION": "PLATRE",
    "PLATRE DE MOULAGE GAZELLE": "PLATRE",
}


def get_category(ar_design):
    """Get category for a product design, returns 'UNCATEGORIZED' if not found."""
    return CATEGORY_MAPPING.get(ar_design.upper(), "UNCATEGORIZED")

print(get_category("AGRIGYPSE PP 50 kg"))

def split_results_by_category(df, column_name="AR_DESIGN"):
    """
    Split query results by category based on the AR_DESIGN column.
    
    Args:
        df (pd.DataFrame): Input dataframe with query results
        column_name (str): Column name to categorize by (default: AR_DESIGN)
    
    Returns:
        dict: Dictionary with categories as keys and dataframes as values
    """
    if column_name not in df.columns:
        print(f"Error: Column '{column_name}' not found in dataframe")
        return {}
    
    # Add category column
    df["CATEGORY"] = df[column_name].apply(get_category)
    
    # Split by category
    categorized_data = {category: group.drop(columns=["CATEGORY"]) 
                       for category, group in df.groupby("CATEGORY")}
    
    return categorized_data


# Database connection using SQLAlchemy
engine = create_engine(
    "mssql+pyodbc:///?odbc_connect=DRIVER={ODBC Driver 17 for SQL Server};"
    "SERVER=.\\SQLEXPRESS;DATABASE=Msyspes2;Trusted_Connection=yes;"
)

# Query parameters
today_day = 6
yesterday_day = 5

today = date.today()
yesterday = today - timedelta(days=1)

query = f"""
SELECT *
FROM dbo._VSORTIES
WHERE 
(
    DAY(CONVERT(DATETIME, DATES, 103)) = 5
    AND POSTE IN ('POSTE1', 'POSTE2')
)
OR
(
    DAY(CONVERT(DATETIME, DATES, 103)) = 6
    AND POSTE = 'POSTE3'
)
"""

# Execute query
df = pd.read_sql(query, engine)

# Apply regex-based HTML entity replacement across all columns
df = df.replace(r"&apos;", "'", regex=True)
df = df.replace(r"&quot;", '"', regex=True)
df = df.replace(r"&amp;", "&", regex=True)
df = df.replace(r"&lt;", "<", regex=True)
df = df.replace(r"&gt;", ">", regex=True)
df = df.replace(r"&#39;", "'", regex=True)

# Also apply html.unescape to catch any other entities
for col in df.columns:
    if df[col].dtype == 'object':
        df[col] = df[col].apply(lambda x: html.unescape(str(x)) if pd.notna(x) else x)

# Normalize AR_DESIGN column to uppercase for consistent categorization
df["AR_DESIGN"] = df["AR_DESIGN"].str.upper()

# Split results by category
categorized_results = split_results_by_category(df, column_name="AR_DESIGN")

# Sort categories by size (largest first)
sorted_categories = sorted(categorized_results.items(), 
                          key=lambda x: len(x[1]), 
                          reverse=True)

# Separate UNCATEGORIZED from the rest
categorized = [(cat, df) for cat, df in sorted_categories if cat != "UNCATEGORIZED"]
uncategorized = [(cat, df) for cat, df in sorted_categories if cat == "UNCATEGORIZED"]

# Combine with UNCATEGORIZED at the end
sorted_categories = categorized + uncategorized


# Create Excel file with formatted output
def create_excel_with_categories(sorted_categories, output_file="automation_report.xlsx"):
    """
    Create an Excel file with categorized data, AR_DESIGN and QTE in the middle,
    4 empty rows between tables, Quantité Total on separate row,
    and UNCATEGORIZED at the end without totals.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    bold_font = Font(bold=True)
    
    # Get first category to determine columns
    first_category, first_df = sorted_categories[0]
    all_columns = first_df.columns.tolist()
    
    # Separate AR_DESIGN and QTE from other columns
    other_columns = [col for col in all_columns if col not in ["AR_DESIGN", "QTE"]]
    
    # Find middle position to insert AR_DESIGN and QTE
    middle_pos = len(other_columns) // 2
    
    # Build reordered columns with AR_DESIGN and QTE in the middle
    reordered_columns = other_columns[:middle_pos] + ["AR_DESIGN", "QTE"] + other_columns[middle_pos:]
    
    # Write header row once
    for col_idx, col_name in enumerate(reordered_columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = bold_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    current_row = 2
    
    # Process each category
    for cat_idx, (category, category_df) in enumerate(sorted_categories):
        # Write data rows for this category with reordered columns
        for row_idx, (_, row) in enumerate(category_df.iterrows()):
            for col_idx, col_name in enumerate(reordered_columns, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=row[col_name])
                
                # Center align AR_DESIGN and QTE columns
                if col_name in ["AR_DESIGN", "QTE"]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            
            current_row += 1
        
        # Add Quantité Total row only if NOT UNCATEGORIZED
        if category != "UNCATEGORIZED":
            ar_design_col = reordered_columns.index("AR_DESIGN") + 1
            
            total_label_cell = ws.cell(row=current_row, column=ar_design_col, 
                                        value=f"Quantité Total {category}")
            total_label_cell.font = bold_font
            total_label_cell.fill = yellow_fill
            total_label_cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # Add sum of QTE if it exists
            if "QTE" in reordered_columns:
                qte_col_idx = reordered_columns.index("QTE") + 1
                total_qty = category_df["QTE"].sum()
                total_qty_cell = ws.cell(row=current_row, column=qte_col_idx, value=total_qty)
                total_qty_cell.font = bold_font
                total_qty_cell.fill = yellow_fill
                total_qty_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            current_row += 1
            
            # Add 4 empty rows between categories (not after UNCATEGORIZED at the end)
            if cat_idx < len(sorted_categories) - 1:
                current_row += 4
        else:
            # For UNCATEGORIZED, just add one empty row between if it's not the last
            if cat_idx < len(sorted_categories) - 1:
                current_row += 1
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save file
    wb.save(output_file)
    print(f"\nExcel file created successfully: {output_file}")
    return output_file

# Create the Excel file
output_file = create_excel_with_categories(sorted_categories, output_file=OUTPUT_FILE)
print(f"Report saved to: {output_file}")

