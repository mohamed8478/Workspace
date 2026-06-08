from datetime import datetime, date, timedelta
import pyodbc
import pandas as pd
from sqlalchemy import create_engine
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import os
import html


# Output configuration
OUTPUT_FOLDER = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(OUTPUT_FOLDER, "Sortie Plâtre par Camion.xlsx")


# Category mapping for product classification
CATEGORY_MAPPING = {
    "SULFATE DE CALCIUM HEMIHYDRATE EN BIG BA": "AGRIGYPSE",
    "MORTIER": "MORTIER",
    "PLATRE GAZELLE EXTRA FIN EN PO": "PLATRE",
    "AGRIGYPSE 0-20 VRAC": "AGRIGYPSE",
    "AGRIGYPSE BIG BAG": "AGRIGYPSE",
    "AGRIGYPSE PP 50 KG": "AGRIGYPSE",
    "AGRIGYPSE EN VRAC": "AGRIGYPSE",
    "AGRIGYPSE PP": "AGRIGYPSE",
    "GYPSE BROYE AGRIGYPSE SAC 70KG": "AGRIGYPSE",
    "SULFATE DE CALCIUM": "AGRIGYPSE",
    "PLATRE EN VRAC": "AGRIGYPSE",
    "GYPSE BROYE": "AGRIGYPSE",
    "DALLE TOURBILLON": "DALLES",
    "DALLE SOLEIL SANS STRUCTURE": "DALLES",
    "DALLE SABLE SANS STRUCTURE": "DALLES",
    "DALLE MEDITERRANE": "DALLES",
    "DALLE HELICE": "DALLES",
    "DALLE FISSURE SANS STRUCTURE": "DALLES",
    "DALLE TOURBILLON DEPART SAFI AVEC STRUCTURE": "DALLES",
    "DALLE TOURBILLON DEPART SAFI SANS STRUCTURE": "DALLES",
    "DALLE SOLEIL DEPART SAFI AVEC STRUCTURE": "DALLES",
    "DALLE SOLEIL DEPART SAFI SANS STRUCTURE": "DALLES",
    "DALLE SABLE DEPART SAFI AVEC STRUCTURE": "DALLES",
    "DALLE SABLE DEPART SAFI SANS STRUCTURE": "DALLES",
    "DALLE PERFORE ROND DEPART SAFI SANS STRUCTURE": "DALLES",
    "DALLE PERFORE CARRE DEPART SAFI AVEC STRUCTURE": "DALLES",
    "DALLE PERFORE CARRE DEPART SAFI SANS STRUCTURE": "DALLES",
    "DALLE MEDITERRANE DEPART SAFI AVEC STRUCTURE": "DALLES",
    "DALLE MEDITERRANE DEPART SAFI SANS STRUCTURE": "DALLES",
    "DALLE HELICE DEPART SAFI AVEC STRUCTURE": "DALLES",
    "DALLE HELICE DEPART SAFI SANS STRUCTURE": "DALLES",
    "DALLE FISSURE DEPART SAFI AVEC STRUCTURE": "DALLES",
    "DALLE FISSURE DEPART SAFI SANS STRUCTURE": "DALLES",
    "AGRIGYPSE PP 50 KG EXPORT": "EXPORT",
    "WHITE GYPSUM ROCK 200-400MM": "EXPORT",
    "PLATRE EXPORT EXTRA FIN GAZELLE": "EXPORT",
    "PLATRE PGC EXPORT NEJMA": "EXPORT",
    "PLATRE EXPORT EXTRA FIN NEJMA": "EXPORT",
    "PLATRE MOULAGE GAZELLE EXPORT": "EXPORT",
    "GYPSE BLANC ROCHE EXPORT BI": "EXPORT",
    "PLATRE PGC EXPORT SAVANE": "EXPORT",
    "PLATRE PGC EXPORT GAZELLE": "EXPORT",
    "GYPSE BLANC CONCASSE EXPORT BI": "EXPORT",
    "PLATRE EXPORT SPECIAL THE DOVE": "EXPORT",
    "PLATRE EXPORT SPECIAL \"THE DOVE\"": "EXPORT",
    "GYPSE BLANC ROCHE EXPORT BIG BAG": "EXPORT",
    "GYPSE BLANC CONCASSE EXPORT BIG BAG": "EXPORT",
    "PLATRE EXPORT MOULAGE SAVANE": "EXPORT",
    "PLATRE EXPORT EN SAC EN BIG BAG": "EXPORT",
    "PLATRE EXPORT SPECIAL GAZELLE": "EXPORT",
    "PLATRE PGC EXPORT GAZELLE_": "EXPORT",
    "PLATRE EXPORT": "EXPORT",
    "ENDUIT JOINT EJ4": "MORTIER",
    "ENDUIT DE PEINTURE": "MORTIER",
    "ENDUIT EJ8": "MORTIER",
    "TALOCHE S9AF VERT PP": "MORTIER",
    "MORTIER DE PLATRE MPREMIUM PP": "MORTIER",
    "MORTIER DE PLATRE GAZELLE M.P.": "MORTIER",
    "MORTIER DE FINITION TOP  FINO P": "MORTIER",
    "MORTIER DE PLATRE GAZELLE TALO": "MORTIER",
    "ENDUIT EJ8 EN PALETTE": "MORTIER",
    "ENDUIT DE PEINTURE ECHANTILLON": "MORTIER",
    "ENDUIT JOINT EJ8 ECHANTILLON": "MORTIER",
    "ENDUIT JOINT EJ4 ech": "MORTIER",
    "MORTIER DE PLATRE MPREMIUM": "MORTIER",
    "ENDUIT DE FINITION": "MORTIER",
    "MORTIER DE PLATRE GAZELLE M.P.E 80 S SUPER": "MORTIER",
    "MORTIER DE PLATRE GAZELLE M.P.E 80 L": "MORTIER",
    "MORTIER EN PALETTE": "MORTIER",
    "MORTIER DE PLATRE GAZELLE TALOCHE 60": "MORTIER",
    "MORTIER MPE 80 BLEU PP": "MORTIER",
    "SPECIAL NEJMA": "PLATRE",
    "SPECIAL DOVE": "PLATRE",
    "SAVANE": "PLATRE",
    "PLATRE MOULAGE SPECIAL POLYPRO": "PLATRE",
    "PLATRE DE CONSTRUCTION ROUGE P": "PLATRE",
    "MOULDING NEJMA": "PLATRE",
    "PLATRE MOULAGE GAZELLE": "PLATRE",
    "EXTRA FIN INDUSTRIEL ECHANTILLON": "PLATRE",
    "PLATRE MOULAGE SPECIAL POLYPROPYLENE": "PLATRE",
    "PLATRE DE MOULAGE SPECIAL POLYPROPYLENE": "PLATRE",
    "PLATRE NEJMA EN POLYPROPYLENE": "PLATRE",
    "PLATRE DE CONSTRUCTION EN POLYPROPYLENE ROUGE": "PLATRE",
    "PLATRE DE CONSTRUCTION EN POLYPROPYLENE BLEU": "PLATRE",
    "PLATRE ( 2 TONNES)": "PLATRE",
    "PLATRE EN SACS POLYPROPYLENE ROUGE": "PLATRE",
    "PLATRE EN STOCK": "PLATRE",
    "PLATRE POLYPROPYLENE EN PALETTE": "PLATRE",
    "PLATRE MOULLAGE GAZELLE EN VRAC": "PLATRE",
    "PLATRE EN BIG BAG": "PLATRE",
    "PLATRE DE MOULAGE SPECIAL": "PLATRE",
    "PLATRE NEJMA": "PLATRE",
    "PLATRE DE CONSTRUCTION": "PLATRE",
    "PLATRE DE MOULAGE GAZELLE": "PLATRE",
}


def get_category(product_name):
    """Get category for a product name, returns 'UNCATEGORIZED' if not found."""
    return CATEGORY_MAPPING.get(product_name.upper(), "UNCATEGORIZED")


def split_results_by_category(df, column_name="NOMPROD"):
    """Split query results by category based on the product name column."""
    if column_name not in df.columns:
        print(f"Error: Column '{column_name}' not found in dataframe")
        return {}
    
    df["CATEGORY"] = df[column_name].apply(get_category)
    categorized_data = {category: group.drop(columns=["CATEGORY"]) 
                       for category, group in df.groupby("CATEGORY")}
    return categorized_data


# Database connection using SQLAlchemy
engine = create_engine(
    "mssql+pyodbc:///?odbc_connect=DRIVER={ODBC Driver 17 for SQL Server};"
    "SERVER=.\\SQLEXPRESS;Database=Msyspes2;Trusted_Connection=yes;"
)

query = f"""
SELECT *
FROM dbo._VSPSORTIE
WHERE 
(
    DAY(CONVERT(DATETIME, DATEPP, 103)) = 5
    AND POSTE IN ('POSTE1', 'POSTE2')
)
OR
(
    DAY(CONVERT(DATETIME, DATEPP, 103)) = 6
    AND POSTE = 'POSTE3'
)
"""

# Execute query
df = pd.read_sql(query, engine)

# Apply regex-based HTML entity replacement across all columns
df = df.replace(r"&apos;", "'", regex=True)
df = df.replace(r"&quot;", '"', regex=True)
df = df.replace(r"&amp;", "&", regex=True)
df = df.replace(r"&lt;", "<", regex=True)
df = df.replace(r"&gt;", ">", regex=True)
df = df.replace(r"&#39;", "'", regex=True)

# Also apply html.unescape to catch any other entities
for col in df.columns:
    if df[col].dtype == 'object':
        df[col] = df[col].apply(lambda x: html.unescape(str(x)) if pd.notna(x) else x)

# Function to combine date and time columns
def combine_datetime(date_str, time_str):
    """Combine date and time strings into a datetime object"""
    try:
        datetime_str = f"{date_str} {time_str}"
        return pd.to_datetime(datetime_str, format='%d/%m/%Y %H:%M')
    except:
        return pd.NaT

# Calculate ATTENT: (DATESMS + HEURESMS) - (DATEPP + HEUREPP)
df['SMS_DateTime'] = df.apply(lambda row: combine_datetime(row['DATESMS'], row['HEURESMS']), axis=1)
df['PP_DateTime'] = df.apply(lambda row: combine_datetime(row['DATEPP'], row['HEUREPP']), axis=1)
df['ATTENT'] = (df['SMS_DateTime'] - df['PP_DateTime']).dt.total_seconds() / 3600  # Convert to hours
df['ATTENT'] = df['ATTENT'].apply(lambda x: 0 if x < 0 else x)

# Calculate Temps de Chargement: (DATESP + HEURESP) - (DATECC + HEURECC)
df['SP_DateTime'] = df.apply(lambda row: combine_datetime(row['DATESP'], row['HEURESP']), axis=1)
df['CC_DateTime'] = df.apply(lambda row: combine_datetime(row['DATECC'], row['HEURECC']), axis=1)
df['Temps de Chargement'] = (df['SP_DateTime'] - df['CC_DateTime']).dt.total_seconds() / 3600  # Convert to hours
df['Temps de Chargement'] = df['Temps de Chargement'].apply(lambda x: 0 if x < 0 else x)

# Convert hours back to timedelta for display as HH:MM:SS format
df['ATTENT_timedelta'] = df['ATTENT'].apply(lambda x: timedelta(hours=x))
df['Temps de Chargement_timedelta'] = df['Temps de Chargement'].apply(lambda x: timedelta(hours=x))

# Drop temporary datetime columns and keep only timedelta versions
df = df.drop(['SMS_DateTime', 'PP_DateTime', 'SP_DateTime', 'CC_DateTime', 'ATTENT', 'Temps de Chargement'], axis=1)
df = df.rename(columns={'ATTENT_timedelta': 'ATTENT', 'Temps de Chargement_timedelta': 'Temps de Chargement'})

# Reorder columns to place ATTENT and Temps de Chargement after X3ON
all_cols = df.columns.tolist()
# Remove the time columns from their current position
time_cols = ['ATTENT', 'Temps de Chargement']
other_cols = [col for col in all_cols if col not in time_cols]

# Find X3ON position and insert time columns after it
if 'X3ON' in other_cols:
    x3on_idx = other_cols.index('X3ON')
    # Insert time columns right after X3ON
    new_col_order = other_cols[:x3on_idx+1] + time_cols + other_cols[x3on_idx+1:]
else:
    # If X3ON doesn't exist, just put time columns at the end
    new_col_order = other_cols + time_cols

df = df[new_col_order]

# Normalize NOMPROD column to uppercase for consistent categorization
df["NOMPROD"] = df["NOMPROD"].str.upper()

# Split results by category
categorized_results = split_results_by_category(df, column_name="NOMPROD")

# Sort categories by size (largest first)
sorted_categories = sorted(categorized_results.items(), 
                          key=lambda x: len(x[1]), 
                          reverse=True)

# Separate UNCATEGORIZED from the rest
categorized = [(cat, df) for cat, df in sorted_categories if cat != "UNCATEGORIZED"]
uncategorized = [(cat, df) for cat, df in sorted_categories if cat == "UNCATEGORIZED"]

sorted_categories = categorized + uncategorized


def create_excel_with_categories(sorted_categories, output_file="automation_report.xlsx"):
    """Create an Excel file with categorized data with time format display and summary rows"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    bold_font = Font(bold=True)
    
    first_category, first_df = sorted_categories[0]
    all_columns = first_df.columns.tolist()
    
    # Write header row
    for col_idx, col_name in enumerate(all_columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = bold_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    current_row = 2
    
    # Process each category
    for cat_idx, (category, category_df) in enumerate(sorted_categories):
        # Write data rows for this category
        for row_idx, (_, row) in enumerate(category_df.iterrows()):
            for col_idx, col_name in enumerate(all_columns, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=row[col_name])
                
                # Format time columns with HH:MM:SS format
                if col_name in ['ATTENT', 'Temps de Chargement']:
                    cell.number_format = '[h]:mm:ss'
            
            current_row += 1
        
        # Add summary row with Taux d'Attente [category] and averages
        if category != "UNCATEGORIZED":
            # Get the column right before ATTENT for the label
            if "ATTENT" in all_columns:
                attent_idx = all_columns.index("ATTENT")
                label_col = attent_idx  # Column right before ATTENT
                
                # Label cell
                label_cell = ws.cell(row=current_row, column=label_col, 
                                    value=f"Taux d'Attente {category}")
                label_cell.font = bold_font
                label_cell.fill = yellow_fill
                label_cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # Calculate and add averages
                attent_col = all_columns.index("ATTENT") + 1
                avg_attent = category_df["ATTENT"].mean()
                attent_cell = ws.cell(row=current_row, column=attent_col, value=avg_attent)
                attent_cell.font = bold_font
                attent_cell.fill = yellow_fill
                attent_cell.number_format = '[h]:mm:ss'
                attent_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            if "Temps de Chargement" in all_columns:
                temps_col = all_columns.index("Temps de Chargement") + 1
                avg_temps = category_df["Temps de Chargement"].mean()
                temps_cell = ws.cell(row=current_row, column=temps_col, value=avg_temps)
                temps_cell.font = bold_font
                temps_cell.fill = yellow_fill
                temps_cell.number_format = '[h]:mm:ss'
                temps_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            current_row += 1
            
            # Add 4 empty rows between categories
            if cat_idx < len(sorted_categories) - 1:
                current_row += 4
        else:
            # For UNCATEGORIZED, just add one empty row
            if cat_idx < len(sorted_categories) - 1:
                current_row += 1
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_file)
    print(f"\nExcel file created successfully: {output_file}")
    return output_file


print("=" * 70)
print("SORTIE PLATRE PAR CAMION - REPORT GENERATION")
print("=" * 70)
print("\nDatabase connection successful!")
print(f"Total rows retrieved: {len(df)}")
print(f"Total categories found: {len(sorted_categories)}")
print("\nCategories breakdown:")
for category, category_df in sorted_categories:
    # Convert timedelta back to hours for display
    avg_attent_td = category_df["ATTENT"].mean()
    avg_temps_td = category_df["Temps de Chargement"].mean()
    
    avg_attent_hours = avg_attent_td.total_seconds() / 3600 if pd.notna(avg_attent_td) else 0
    avg_temps_hours = avg_temps_td.total_seconds() / 3600 if pd.notna(avg_temps_td) else 0
    
    print(f"  - {category}: {len(category_df)} rows | Avg Attente: {avg_attent_hours:.2f}h | Avg Chargement: {avg_temps_hours:.2f}h")
print("=" * 70)

output_file = create_excel_with_categories(sorted_categories, output_file=OUTPUT_FILE)
print(f"Report saved to: {output_file}")
print("=" * 70)

